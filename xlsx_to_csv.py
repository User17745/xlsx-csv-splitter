#!/usr/bin/env python3
"""
XLSX to CSV Splitter
--------------------
Split an Excel file with multiple worksheets into separate CSV files.
Optimized for LLM ingestion with clean output and proper encoding.

Usage:
    python xlsx_to_csv.py <input_file.xlsx>
    python xlsx_to_csv.py <input_file.xlsx> --output-dir ./csv_output
    python xlsx_to_csv.py <input_file.xlsx> --sheet-names "Sheet1,Sheet2"

Requirements:
    pip install openpyxl
"""

import argparse
import csv
import os
import sys
from pathlib import Path
from typing import Optional


def sanitize_filename(name: str) -> str:
    """Sanitize worksheet name for use as filename."""
    # Replace invalid filename characters with underscores
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '_')
    return name.strip()


def get_worksheet_data(ws, max_empty_rows: int = 3) -> list[list[str]]:
    """
    Extract data from worksheet, handling merged cells and stopping at
    consecutive empty rows (LLM ingestion optimization).
    """
    data = []
    empty_row_count = 0

    for row in ws.iter_rows(values_only=True):
        # Convert all values to strings, handling None
        row_data = [str(v) if v is not None else '' for v in row]

        # Check if row is empty
        if all(not cell.strip() for cell in row_data):
            empty_row_count += 1
            if empty_row_count >= max_empty_rows:
                break
            continue

        empty_row_count = 0
        data.append(row_data)

    return data


def write_csv(data: list[list[str]], filepath: Path) -> None:
    """Write data to CSV file with UTF-8 encoding."""
    if not data:
        return

    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerows(data)


def split_xlsx_to_csv(
    input_file: Path,
    output_dir: Optional[Path] = None,
    sheet_names: Optional[list[str]] = None,
    include_index: bool = False,
    max_empty_rows: int = 3,
) -> dict[str, Path]:
    """
    Split XLSX file into multiple CSV files.

    Args:
        input_file: Path to input XLSX file
        output_dir: Directory for output CSVs (default: same as input)
        sheet_names: List of sheet names to export (default: all)
        include_index: Add row index as first column
        max_empty_rows: Stop after this many consecutive empty rows

    Returns:
        Dictionary mapping sheet names to output file paths
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    # Set output directory
    if output_dir is None:
        output_dir = input_file.parent / f"{input_file.stem}_csv"
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load workbook
    print(f"Loading: {input_file}")
    wb = load_workbook(input_file, data_only=True, read_only=True)

    # Determine which sheets to process
    sheets_to_process = sheet_names if sheet_names else wb.sheetnames
    results = {}

    for sheet_name in sheets_to_process:
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found, skipping.")
            continue

        ws = wb[sheet_name]
        data = get_worksheet_data(ws, max_empty_rows)

        if not data:
            print(f"Warning: Sheet '{sheet_name}' is empty, skipping.")
            continue

        # Add row index if requested
        if include_index:
            for i, row in enumerate(data):
                data[i] = [str(i)] + row

        # Create output filename
        safe_name = sanitize_filename(sheet_name)
        output_file = output_dir / f"{safe_name}.csv"

        write_csv(data, output_file)
        results[sheet_name] = output_file
        print(f"  Exported: {sheet_name} -> {output_file} ({len(data)} rows)")

    wb.close()
    return results


def main():
    parser = argparse_parser = argparse.ArgumentParser(
        description="Split XLSX file into multiple CSV files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s data.xlsx
  %(prog)s data.xlsx --output-dir ./output
  %(prog)s data.xlsx --sheet-names "Sheet1,Sheet2" --include-index
        """,
    )
    parser.add_argument('input', type=Path, help='Input XLSX file path')
    parser.add_argument(
        '-o', '--output-dir',
        type=Path,
        help='Output directory for CSV files (default: <input_name>_csv)',
    )
    parser.add_argument(
        '-s', '--sheet-names',
        help='Comma-separated list of sheet names to export (default: all)',
    )
    parser.add_argument(
        '-i', '--include-index',
        action='store_true',
        help='Add row index as first column',
    )
    parser.add_argument(
        '--max-empty-rows',
        type=int,
        default=3,
        help='Stop after N consecutive empty rows (default: 3)',
    )

    args = parser.parse_args()

    sheet_names = None
    if args.sheet_names:
        sheet_names = [s.strip() for s in args.sheet_names.split(',')]

    results = split_xlsx_to_csv(
        input_file=args.input,
        output_dir=args.output_dir,
        sheet_names=sheet_names,
        include_index=args.include_index,
        max_empty_rows=args.max_empty_rows,
    )

    if results:
        print(f"\nSuccessfully exported {len(results)} sheet(s) to CSV.")
    else:
        print("\nNo sheets were exported.")
        sys.exit(1)


if __name__ == '__main__':
    main()
